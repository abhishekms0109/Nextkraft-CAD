"""
AutoCAD Tower Parking – driven by Excel (level_data.xlsm).

Excel layout:
  - B/C: PARAMETER / VALUES (main table; also used for AutoCAD table and levels).
  - F/G: MANUAL ENTRY – overrides for SEDAN_LEVEL and SUV_LEVEL.
  - I/J: COMMAND / VALUE – optional drawing commands (TITLE_TEXT, NOTE_TEXT, LABEL_TEXT_HEIGHT, DIMENSION_OFFSET_X, etc.).

To add new drawing commands: add a row in I/J and a handler in COMMAND_HANDLERS (see EXCEL_COMMANDS_README.md).
"""
from openpyxl import load_workbook
from pyautocad import Autocad, APoint

# ---------- 1. Read Excel Data (unified: main table B/C + MANUAL ENTRY F/G) ----------
def get_all_excel_params(file_path, param_sheet_name=None):
    """Read full parameter table from B/C. Apply MANUAL ENTRY (F/G) overrides for SEDAN_LEVEL, SUV_LEVEL."""
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb[param_sheet_name] if param_sheet_name and param_sheet_name in wb.sheetnames else wb.active
    params = {}
    row = 2
    while True:
        param = ws[f"B{row}"].value
        value = ws[f"C{row}"].value
        if param is None or str(param).strip() == "":
            break
        key = str(param).strip().upper()
        if value is not None:
            params[key] = int(value) if isinstance(value, (int, float)) and not isinstance(value, bool) else str(value)
        else:
            params[key] = ""
        row += 1
    # MANUAL ENTRY override (columns F/G)
    row = 2
    while True:
        cmd = ws[f"F{row}"].value
        val = ws[f"G{row}"].value
        if cmd is None or str(cmd).strip() == "":
            break
        key = str(cmd).strip().upper()
        if key in ("SEDAN_LEVEL", "SUV_LEVEL") and val is not None:
            try:
                params[key] = int(val)
            except (ValueError, TypeError):
                params[key] = val
        row += 1
    return params

def read_levels_from_excel(file_path, use_output_sheet=True):
    """Levels for blocks: main sheet B/C + MANUAL ENTRY F/G, then merge OUTPUT sheet if present."""
    params = get_all_excel_params(file_path)
    if use_output_sheet:
        try:
            wb = load_workbook(filename=file_path, data_only=True)
            if "OUTPUT" in wb.sheetnames:
                sheet = wb["OUTPUT"]
                for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
                    if row[1] and row[2] and isinstance(row[2], (int, float)):
                        params[str(row[1]).strip().upper()] = int(row[2])
        except Exception:
            pass
    return {
        "SUV_LEVEL": int(params.get("SUV_LEVEL", 0)),
        "SEDAN_LEVEL": int(params.get("SEDAN_LEVEL", 0)),
        "SUV_SLOT": int(params.get("SUV_SLOT", 2100)),
        "SEDAN_SLOT": int(params.get("SEDAN_SLOT", 1800)),
    }

def get_excel_table_data(filename, param_sheet_name=None):
    """Table rows for AutoCAD table: PARAMETER (B) and VALUES (C)."""
    wb = load_workbook(filename=filename, data_only=True)
    ws = wb[param_sheet_name] if param_sheet_name and param_sheet_name in wb.sheetnames else wb.active
    data = []
    row = 2
    while True:
        param = ws[f"B{row}"].value
        value = ws[f"C{row}"].value
        if param is None or str(param).strip() == "":
            break
        data.append((str(param), "" if value is None else str(value)))
        row += 1
    return data

def get_excel_commands(file_path, param_sheet_name=None):
    """Optional COMMAND/VALUE list from columns I/J (same sheet). Add custom drawing commands here."""
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb[param_sheet_name] if param_sheet_name and param_sheet_name in wb.sheetnames else wb.active
    commands = []
    for row in range(2, 200):
        cmd = ws[f"I{row}"].value
        val = ws[f"J{row}"].value
        if cmd is None or str(cmd).strip() == "":
            continue
        commands.append((str(cmd).strip().upper(), val))
    return commands

# ---------- 2. Insert Parking Stack Blocks ----------
from openpyxl import load_workbook
from pyautocad import Autocad, APoint

# ==== CONFIGURATION ====
SUV_BLOCK             = "SUV_BLOCK"
SEDAN_BLOCK           = "SEDAN_BLOCK"
GROUND_BLOCK          = "GROUND_LEVEL_BLOCK"
MACHINE_ROOM_BLOCK    = "MACHINE_ROOM_BLOCK"

SUV_SPACING           = 2100    # mm between each SUV level
SEDAN_SPACING         = 1800    # mm between each Sedan level
GROUND_BLOCK_HEIGHT   = 6913    # mm from Y=0 to top of ground block
SUV_BLOCK_HEIGHT      = 2100    # mm actual height of SUV block
SEDAN_BLOCK_HEIGHT    = 1800    # mm actual height of Sedan block

EXCEL_PATH            = "level_data.xlsm"  # or .xlsx if you changed it

def insert_blocks(levels, acad):
    suv_count   = levels.get("SUV_LEVEL", 0)
    sedan_count = levels.get("SEDAN_LEVEL", 0)

    # Get dynamic slot heights (default to original if missing)
    suv_height = levels.get("SUV_SLOT", SUV_SPACING)
    sedan_height = levels.get("SEDAN_SLOT", SEDAN_SPACING)

    # Calculate Y-axis scale factors
    # Base height assumed to be the default spacing (2100 for SUV, 1800 for Sedan)
    suv_scale_y = suv_height / SUV_BLOCK_HEIGHT
    sedan_scale_y = sedan_height / SEDAN_BLOCK_HEIGHT
    
    print(f"📏 Scaling SUV: {suv_scale_y:.4f} (Height: {suv_height})")
    print(f"📏 Scaling Sedan: {sedan_scale_y:.4f} (Height: {sedan_height})")

    # 1️⃣ Ground
    acad.model.InsertBlock(APoint(0, 0), GROUND_BLOCK, 1,1,1,0)
    print("✅ Ground at Y=0")

    # start stacking at top of ground
    y = GROUND_BLOCK_HEIGHT

    # 2️⃣ SUV stack
    for i in range(suv_count):
        # Insert with Y-scaling
        acad.model.InsertBlock(APoint(0, y), SUV_BLOCK, 1, suv_scale_y, 1, 0)
        print(f"✅ SUV #{i+1} at Y={y}")
        y += suv_height

    # 🆕 SUV-ONLY CASE: if no Sedans, place Machine Room right here and exit
    if sedan_count == 0 and suv_count > 0:
        machine_y = y  - SUV_BLOCK_HEIGHT # Stick to top of last block (visually checking logic might be needed here if heights differ significantly)
        # For Machine Room, we do not scale it, usually sits on top.
        # Adjusted logic: y is currently at (top of last SUV).
        # We want machine room base at top of last SUV.
        # Original code: machine_y = y - SUV_BLOCK_HEIGHT. 
        # Wait, if y increments AFTER insertion, y is at position of NEXT block.
        # So top of last block is y (current) - suv_height? No.
        # Original loop: insert at y, then y += spacing.
        # So after loop, y is at (last_y + spacing). Top of last block is approximately y if spacing == height.
        # Let's trust original logic for position but use current y - actual height inserted?
        # Actually, if we just place it at 'y' it might float if there's a gap?
        # In original code: machine_y = y - SUV_BLOCK_HEIGHT.
        # If SUV_BLOCK_HEIGHT is 2100 and spacing is 2100.
        # New code: y has incremented by suv_height.
        # So we should subtract suv_height to get to the start of the last block? 
        # No, we want it ON TOP.
        # Let's look at previous loop: Insert at 0, Y=6913. Next Y = 6913+2100.
        # Top of first block is 6913+2100.
        # So 'y' IS the top of the stack.
        # Original code: machine_y = y - SUV_BLOCK_HEIGHT. 
        # This implies the Machine Room insertion point is NOT at the bottom of the machine room, 
        # OR the SUV block insertion point is at the bottom, and we are placing machine room overlapping the last block?
        # Let's assume the previous logic was correct for the specific block definitions.
        # Use suv_height to be consistent with the scaling.
        machine_y = y - suv_height
        acad.model.InsertBlock(APoint(0, machine_y), MACHINE_ROOM_BLOCK, 1,1,1,0)
        print(f"✅ (SUV-only) Machine Room at Y={machine_y}")
        return

    # 3️⃣ Sedan stack (if any)
    for i in range(sedan_count):
        acad.model.InsertBlock(APoint(0, y), SEDAN_BLOCK, 1, sedan_scale_y, 1, 0)
        print(f"✅ Sedan #{i+1} at Y={y}")
        y += sedan_height

    # 4️⃣ Machine Room on top
    # Use the height of the last placed block type to backtrack if needed
    last_height = sedan_height if sedan_count > 0 else suv_height
    machine_y = y - last_height
    acad.model.InsertBlock(APoint(0, machine_y), MACHINE_ROOM_BLOCK, 1,1,1,0)
    print(f"✅ Machine Room at Y={machine_y}")

if __name__ == "__main__":
    acad   = Autocad(create_if_not_exists=True)
    levels = read_levels_from_excel(EXCEL_PATH)
    print("🚗 Levels →", levels)
    insert_blocks(levels, acad)
    print("🏁 Done.")


# ---------- 3. Place PARKING_BLOCK on PLACE_HOLDER ----------
def place_parking_blocks_on_placeholders(acad):
    PLACE_HOLDER = "PLACE_HOLDER"
    PARKING_BLOCK = "PARKING_BLOCK"
    
    # Scaling factor for Parking Block (6700mm -> 7000mm)
    # Target / Base
    scale_x = 7000.0 / 6700.0
    
    count = 0

    for obj in acad.iter_objects("AcDbBlockReference"):
        try:
            if obj.EffectiveName.upper() == PLACE_HOLDER:
                ins_pt = APoint(obj.InsertionPoint)
                # Apply X scaling
                acad.model.InsertBlock(ins_pt, PARKING_BLOCK, scale_x, 1, 1, 0)
                print(f"📌 PARKING_BLOCK placed at {ins_pt} with Scale X={scale_x:.4f}")
                count += 1
        except Exception as e:
            print(f"⚠️ Skipped one object due to error: {e}")

    print(f"🆗 {count} PARKING_BLOCK(s) inserted on PLACE_HOLDER block(s).")

# ---------- 4. Add Floor Labels and Total Height Dimension ----------
def add_floor_labels_and_dimension(levels, acad, params=None):
    params = params or {}
    suv_num_blocks = levels.get("SUV_LEVEL", 0)
    sedan_num_blocks = levels.get("SEDAN_LEVEL", 0)
    
    # Dynamic Spacing
    suv_spacing = levels.get("SUV_SLOT", SUV_SPACING)
    sedan_spacing = levels.get("SEDAN_SLOT", SEDAN_SPACING)
    
    SUV_START_X = 11600
    SUV_START_Y = 7400
    LABEL_LINE_LENGTH = 2500
    LABEL_TEXT_HEIGHT = params.get("_LABEL_TEXT_HEIGHT", 400)

    # SUV Labels
    for i in range(suv_num_blocks):
        y = SUV_START_Y + i * suv_spacing
        label = f"{i+1}F"
        acad.model.AddText(label, APoint(SUV_START_X, y), LABEL_TEXT_HEIGHT)
        ln_y = y - LABEL_TEXT_HEIGHT / 2 - 100
        x1 = SUV_START_X - LABEL_LINE_LENGTH / 2
        x2 = SUV_START_X + LABEL_LINE_LENGTH / 2
        acad.model.AddLine(APoint(x1, ln_y), APoint(x2, ln_y))
        print(f"Added {label} and line at X={SUV_START_X}, Y={y}")

    # Sedan Labels (continue numbering)
    SEDAN_START_Y = SUV_START_Y + (suv_spacing * suv_num_blocks)
    for i in range(sedan_num_blocks):
        y = SEDAN_START_Y + i * sedan_spacing
        label = f"{suv_num_blocks + i + 1}F"
        acad.model.AddText(label, APoint(SUV_START_X, y), LABEL_TEXT_HEIGHT)
        ln_y = y - LABEL_TEXT_HEIGHT / 2 - 100
        x1 = SUV_START_X - LABEL_LINE_LENGTH / 2
        x2 = SUV_START_X + LABEL_LINE_LENGTH / 2
        acad.model.AddLine(APoint(x1, ln_y), APoint(x2, ln_y))
        print(f"Added {label} and line at X={SUV_START_X}, Y={y}")

    # Total Dimension
    DIM_Y_START = 3912.51
    DIM_Y_END = DIM_Y_START + 3000 + (suv_spacing * suv_num_blocks) + (sedan_spacing * sedan_num_blocks) + 4200
    dim_offset_x = params.get("_DIMENSION_OFFSET_X", -3150)
    p1 = APoint(0, DIM_Y_START)
    p2 = APoint(0, DIM_Y_END)
    dim_line_point = APoint(dim_offset_x, (DIM_Y_START + DIM_Y_END) / 2)
    acad.model.AddDimAligned(p1, p2, dim_line_point)
    print(f"📏 Added total height dimension from Y={DIM_Y_START} to Y={DIM_Y_END} at X={dim_offset_x}")

# ---------- 5. Update AutoCAD Table ----------
def update_acad_table(excel_data, acad):
    table_found = False
    for obj in acad.iter_objects('AcDbTable'):
        table_found = True
        print("✅ Found target table, updating...")
        nrows = obj.Rows
        ncols = obj.Columns
        for i, (param, val) in enumerate(excel_data):
            if i + 1 >= nrows:
                break  # Prevent overflow
            obj.SetText(i + 1, 0, param)  # Left column
            obj.SetText(i + 1, 1, val)    # Right column
        print("✅ Table updated from Excel.")
        break  # Only update the first found table
    if not table_found:
        print("❌ No AutoCAD table found in this drawing.")


# ---------- 6. Extensible Excel-driven drawing commands (I/J columns) ----------
def _cmd_title_text(value, acad, params):
    """Place title text at top-left. COMMAND: TITLE_TEXT, VALUE: your title string."""
    if not value:
        return
    try:
        pt = APoint(500, params.get("TITLE_Y", 45000))
        acad.model.AddText(str(value), pt, params.get("TITLE_HEIGHT", 800))
        print(f"✅ Title text: {value}")
    except Exception as e:
        print(f"⚠️ TITLE_TEXT: {e}")

def _cmd_note_text(value, acad, params):
    """Place note text. COMMAND: NOTE_TEXT, VALUE: text (use | for newline)."""
    if not value:
        return
    try:
        x, y = params.get("NOTE_X", 500), params.get("NOTE_Y", 40000)
        height = params.get("NOTE_HEIGHT", 300)
        for line in str(value).split("|"):
            acad.model.AddText(line, APoint(x, y), height)
            y -= height * 1.2
        print(f"✅ Note text added.")
    except Exception as e:
        print(f"⚠️ NOTE_TEXT: {e}")

def _cmd_label_text_height(value, acad, params):
    """Override floor label text height (mm). COMMAND: LABEL_TEXT_HEIGHT, VALUE: number."""
    try:
        h = float(value)
        params["_LABEL_TEXT_HEIGHT"] = h
        print(f"✅ Label text height override: {h}")
    except (ValueError, TypeError):
        print(f"⚠️ LABEL_TEXT_HEIGHT: invalid value {value}")

def _cmd_dimension_offset_x(value, acad, params):
    """Override dimension line X offset (mm). COMMAND: DIMENSION_OFFSET_X, VALUE: number."""
    try:
        params["_DIMENSION_OFFSET_X"] = float(value)
        print(f"✅ Dimension offset X override: {value}")
    except (ValueError, TypeError):
        print(f"⚠️ DIMENSION_OFFSET_X: invalid value {value}")

# Registry: add new Excel commands here. Key = name in Excel (I column), Value = handler(value, acad, params).
COMMAND_HANDLERS = {
    "TITLE_TEXT": _cmd_title_text,
    "NOTE_TEXT": _cmd_note_text,
    "LABEL_TEXT_HEIGHT": _cmd_label_text_height,
    "DIMENSION_OFFSET_X": _cmd_dimension_offset_x,
}

def run_excel_commands(commands, acad, params):
    """Run optional drawing commands from Excel columns I (COMMAND) and J (VALUE)."""
    for cmd_name, value in commands:
        if cmd_name in COMMAND_HANDLERS:
            try:
                COMMAND_HANDLERS[cmd_name](value, acad, params)
            except Exception as e:
                print(f"⚠️ Command {cmd_name}: {e}")
        else:
            print(f"⚠️ Unknown command: {cmd_name} (add handler in COMMAND_HANDLERS)")


# ---------- MAIN: Run All Steps ----------
if __name__ == "__main__":
    excel_path = EXCEL_PATH
    print("Reading Excel...")
    all_params = get_all_excel_params(excel_path)
    levels = read_levels_from_excel(excel_path)
    excel_data = get_excel_table_data(excel_path)
    commands = get_excel_commands(excel_path)
    print("Levels:", levels)
    print("Table rows:", len(excel_data))
    if commands:
        print("Custom commands:", [c[0] for c in commands])
    acad = Autocad(create_if_not_exists=True)
    print("Connected to AutoCAD:", acad.doc.Name)

    # 1. Insert blocks
    insert_blocks(levels, acad)
    # 2. Place parking blocks
    place_parking_blocks_on_placeholders(acad)
    # 3. Run optional commands from Excel I/J first (sets overrides like LABEL_TEXT_HEIGHT, DIMENSION_OFFSET_X)
    run_excel_commands(commands, acad, all_params)
    # 4. Add labels and dimension (uses all_params, including overrides from commands)
    add_floor_labels_and_dimension(levels, acad, all_params)
    # 5. Update table
    update_acad_table(excel_data, acad)
    print("🚩 All auto-parking proposal steps completed!")
